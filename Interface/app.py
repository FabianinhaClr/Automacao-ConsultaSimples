# app.py
import os
import re
import time
import calendar
from io import BytesIO
from datetime import datetime, date
from dateutil import parser as dparser

import pandas as pd
import requests
from openpyxl import load_workbook
import streamlit as st

# ===========================
# Configuração da página
# ===========================
st.set_page_config(page_title="Consulta Simples Nacional", page_icon="📄", layout="centered")

# ===========================
# Frase de efeito (customize à vontade)
# ===========================
st.markdown(
    """
    <h1 style="text-align:center; margin-bottom:0.2rem;">📄 Consulta de Regime — Simples Nacional</h1>
    <p style="text-align:center; font-size:1.1rem; opacity:0.85;">
        Suba sua planilha e descubra, mês a mês, o regime tributário dos CNPJs em segundos.
    </p>
    """,
    unsafe_allow_html=True
)

# ===========================
# Utilitários (seu código, adaptado)
# ===========================
def clean_cnpj(s: str) -> str:
    return re.sub(r'\D', '', str(s)).zfill(14)

def parse_date_any(s):
    if not s:
        return None
    s = str(s).strip()
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return dparser.parse(s, dayfirst=True).date()
    except Exception:
        return None

def _get_value(item, keys):
    for k in keys:
        if isinstance(item, dict) and k in item and item[k] not in (None, ""):
            return item[k]
    return None

def extract_periods_from_response(resp_json: dict):
    periods = []
    if not resp_json or not isinstance(resp_json, dict):
        return periods

    data = resp_json.get("data")
    data_root = None

    if isinstance(data, list) and data:
        data_root = data[0]
    elif isinstance(data, dict):
        data_root = data
    else:
        data_root = resp_json

    candidate_list_keys = [
        "simples_nacional_periodos_anteriores",
        "simples_nacional_periodos",
        "periodos_simples",
        "simples_periodos",
        "simples_nacional",
        "periodos",
        "permanencia",
        "periodo"
    ]

    for k in candidate_list_keys:
        lst = data_root.get(k) if isinstance(data_root, dict) else None
        if isinstance(lst, list):
            for item in lst:
                if not isinstance(item, dict):
                    continue
                s = _get_value(item, ["inicio_data", "data_inicio", "inicio", "data"])
                e = _get_value(item, ["fim_data", "data_fim", "fim", "data_fim"])
                detalhe = _get_value(item, ["detalhamento", "detalhe", "motivo"]) or ""
                si = parse_date_any(s)
                ei = parse_date_any(e)
                if si:
                    periods.append({"start": si, "end": ei, "detalhe": detalhe})
            if periods:
                return periods

    return periods

def month_date_range(year, month):
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    return first_day, last_day

def is_month_fully_covered(periods, year, month):
    start_month, end_month = month_date_range(year, month)
    hoje = date.today()

    for p in periods:
        si = p.get("start")
        ei = p.get("end")
        detalhe = p.get("detalhe") or ""
        if not si:
            continue
        ei_efetivo = ei if ei else hoje

        if si <= start_month and ei_efetivo >= end_month:
            if year == hoje.year and month == hoje.month and ei is None:
                return True, "Status atual é Simples Nacional."
            else:
                return True, "Permaneceu no Simples Nacional o mês inteiro."

    for p in periods:
        si = p.get("start")
        ei = p.get("end")
        if not si or not ei:
            continue
        if si <= end_month and ei < end_month:
            excl_data = ei.strftime("%Y-%m-%d")
            return False, f"Excluída do Simples Nacional em {excl_data}."

    return False, "Não optante/Nunca esteve no Simples Nacional neste mês."

def read_cnpjs_from_df(df: pd.DataFrame):
    if 'cnpj_part' not in df.columns:
        raise ValueError("A planilha deve conter a coluna 'cnpj_part'.")
    cnpjs = (
        df['cnpj_part']
        .dropna()
        .astype(str)
        .apply(clean_cnpj)
        .unique()
    )
    return cnpjs

def query_infosimples(cnpj, api_url: str, api_key: str, debug: bool = False, log_fn=print):
    try:
        if api_url and api_url.strip():
            args = {"cnpj": cnpj, "token": api_key, "timeout": 300}
            r = requests.post(api_url, data=args, timeout=60)
        else:
            r = requests.get(f"https://www.receitaws.com.br/v1/cnpj/{cnpj}", timeout=60)

        try:
            j = r.json()
        except Exception:
            j = None

        if debug:
            log_fn("=" * 80)
            log_fn(f"[DEBUG] CNPJ PROCESSADO: {cnpj}")
            log_fn("[DEBUG] JSON COMPLETO:")
            import json
            log_fn(json.dumps(j, indent=2, ensure_ascii=False))

        return r.status_code if 'r' in locals() else None, j
    except Exception as e:
        if debug:
            log_fn(f"[DEBUG] Erro na requisição para {cnpj}: {e}")
        return None, None

def process_dataframe(
    df_input: pd.DataFrame,
    api_url: str,
    api_key: str,
    start_year: int = 2020,
    sleep_seconds: float = 0.5,
    debug: bool = False,
    progress_cb=lambda x: None,
    log_fn=lambda *args, **kwargs: None
):
    cnpjs = read_cnpjs_from_df(df_input)
    total = len(cnpjs)
    rows = []
    hoje = date.today()

    for idx, cnpj in enumerate(cnpjs, start=1):
        progress_cb(idx / total if total else 1.0)

        status, resp_json = query_infosimples(cnpj, api_url, api_key, debug, log_fn)
        periods = extract_periods_from_response(resp_json)

        situacao_atual = None
        if resp_json and "data" in resp_json:
            data_field = resp_json["data"]
            if isinstance(data_field, list) and len(data_field) > 0:
                data_item = data_field[0]
            elif isinstance(data_field, dict):
                data_item = data_field
            else:
                data_item = {}

            situacao_atual = (
                data_item.get("simples_nacional_situacao")
                or data_item.get("situacao_simples")
                or data_item.get("situacao")
            )

        texto_situacao = (situacao_atual or "").lower()

        if "optante pelo simples nacional" in texto_situacao:
            m = re.search(r"desde\s+(\d{2}/\d{2}/\d{4})", texto_situacao)
            if m:
                start_date = parse_date_any(m.group(1))
            else:
                start_date = date(hoje.year, 1, 1)

            has_open_period = any(p.get("end") is None for p in periods)
            if not has_open_period:
                periods.append({
                    "start": start_date,
                    "end": None,
                    "detalhe": "Situação Atual: Optante pelo Simples Nacional"
                })

        for year in range(start_year, hoje.year + 1):
            for month in range(1, 13):
                if year == hoje.year and month > hoje.month:
                    continue

                regime, motivo = is_month_fully_covered(periods, year, month)
                regime_str = "Simples Nacional" if regime else "Outro Regime"

                mes_data = date(year, month, 1)
                mes_str = mes_data.strftime("%d/%m/%Y")

                periods_str = "; ".join([
                    f"{p['start']} - {p.get('end', 'até hoje')} [{p.get('detalhe', '')}]"
                    for p in periods
                ])

                rows.append({
                    "CNPJ": cnpj,
                    "MÊS": mes_str,
                    "REGIME": regime_str,
                    "MOTIVO": motivo,
                    "Períodos_detectados": periods_str,
                    "Situacao_Atual": situacao_atual or ""
                })

        time.sleep(sleep_seconds)

    df_result = pd.DataFrame(rows, columns=[
        "CNPJ", "MÊS", "REGIME", "MOTIVO", "Períodos_detectados", "Situacao_Atual"
    ])
    return df_result
def add_sheet_into_excel_bytes(original_file_bytes: bytes, df_to_add: pd.DataFrame, sheet_name="CONSULTA") -> bytes:
    """
    Carrega o Excel, garante que ao menos 1 planilha esteja visível,
    substitui (ou cria) a aba `sheet_name` e devolve os bytes do arquivo atualizado.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(original_file_bytes))

    def visible_count(workbook):
        return sum(1 for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible")

    created_tmp = False
    tmp_name = "_tmp_visible_"

    # Se vamos remover a folha alvo e ela é a única visível, cria uma temporária antes
    if sheet_name in wb.sheetnames:
        if visible_count(wb) <= 1:
            wb.create_sheet(tmp_name)  # visível por padrão
            created_tmp = True
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    # Se por algum motivo não sobrou nenhuma visível (arquivos com todas ocultas), cria temporária
    if visible_count(wb) == 0:
        wb.create_sheet(tmp_name)
        created_tmp = True

    out_buf = BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        writer.book = wb
        writer.sheets = {ws.title: ws for ws in wb.worksheets}

        # Cria/escreve a nova aba CONSULTA
        df_to_add.to_excel(writer, sheet_name=sheet_name, index=False)

        # Define CONSULTA como ativa (opcional)
        try:
            writer.book.active = writer.book.sheetnames.index(sheet_name)
        except Exception:
            pass

        # Remove a temporária somente depois que CONSULTA já existe (há outra visível)
        if created_tmp and tmp_name in writer.book.sheetnames:
            ws_tmp = writer.book[tmp_name]
            writer.book.remove(ws_tmp)

        writer.save()

    out_buf.seek(0)
    return out_buf.getvalue()

# ===========================
# Sidebar (opções)
# ===========================
with st.sidebar:
    st.header("⚙️ Opções")
    api_url = st.text_input("API_URL (opcional)", os.getenv("API_URL", ""))
    api_key = st.text_input("API_KEY (opcional)", os.getenv("API_KEY", ""), type="password")
    start_year = st.number_input("Ano inicial de análise", min_value=2000, max_value=date.today().year, value=2020)
    sleep_seconds = st.number_input("Intervalo entre CNPJs (segundos)", min_value=0.0, max_value=10.0, value=0.5, step=0.1)
    debug = st.checkbox("Ativar DEBUG (log detalhado)", value=True)

# ===========================
# Upload da Planilha
# ===========================
uploaded = st.file_uploader("Envie sua planilha Excel (.xlsx) com a coluna 'cnpj_part'", type=["xlsx"])

# Espaço para logs no modo DEBUG
log_area = st.empty()
def st_log(msg):
    if debug:
        log_area.write(f"```\n{msg}\n```")

# ===========================
# Botão principal
# ===========================
run = st.button("Rodar consulta", type="primary", disabled=(uploaded is None))

if run:
    try:
        if uploaded is None:
            st.warning("Envie um arquivo Excel primeiro.")
            st.stop()

        try:
            df_in = pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"Não foi possível ler o Excel enviado: {e}")
            st.stop()

        if 'cnpj_part' not in df_in.columns:
            st.error("A planilha deve conter a coluna 'cnpj_part'.")
            st.stop()

        progress = st.progress(0, text="Processando CNPJs...")

        def progress_cb(p):
            progress.progress(int(p * 100), text="Processando CNPJs...")

        with st.spinner("Consultando dados e gerando a aba 'CONSULTA'..."):
            df_result = process_dataframe(
                df_input=df_in,
                api_url=api_url,
                api_key=api_key,
                start_year=start_year,
                sleep_seconds=sleep_seconds,
                debug=debug,
                progress_cb=progress_cb,
                log_fn=st_log
            )

            # Gera um Excel novo com a aba CONSULTA inserida no arquivo original
            uploaded.seek(0)
            original_bytes = uploaded.read()
            out_bytes = add_sheet_into_excel_bytes(original_bytes, df_result, sheet_name="CONSULTA")

        st.success("✅ Consulta finalizada! A aba 'CONSULTA' foi gerada.")
        st.caption(f"Total de linhas na CONSULTA: {len(df_result):,}".replace(",", "."))

        st.subheader("Prévia do resultado")
        st.dataframe(df_result.head(50), use_container_width=True)

        st.download_button(
            label="⬇️ Baixar Excel com a aba CONSULTA",
            data=out_bytes,
            file_name="consulta_atualizada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Ocorreu um erro: {e}")
